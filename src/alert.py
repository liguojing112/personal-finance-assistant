"""
超支提醒模块
读取budget.json，与月度实际支出对比，输出超支明细，并在Excel中标记超支单元格
支持默认预算（当月份未在budget.json中定义时使用）
"""
import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 默认预算（当budget.json中没有指定月份时使用）
DEFAULT_BUDGET = {
    "总预算": 1000,
    "居住": 200,
    "餐饮": 300,
    "购物": 200,
    "社交": 100,
    "学习": 50,
    "交通": 50,
    "其他": 100
}

def load_budget(budget_path='config/budget.json'):
    """加载预算配置"""
    if not Path(budget_path).exists():
        print(f"警告：预算文件 {budget_path} 不存在，将仅使用默认预算")
        return {}
    with open(budget_path, 'r', encoding='utf-8') as f:
        budget = json.load(f)
    return budget

def get_budget_for_month(month, budget):
    """
    获取某个月的预算字典，优先使用budget中的定义，否则返回DEFAULT_BUDGET的副本
    """
    if month in budget:
        return budget[month]
    else:
        print(f"月份 {month} 未在budget.json中定义，使用默认预算")
        return DEFAULT_BUDGET.copy()

def check_overbudget(df_expense, budget, month):
    """
    检查某月是否超支（旧版函数，保留兼容性，返回{类别: 超出金额}）
    建议使用get_overbudget_details获取详细信息
    """
    month_budget = get_budget_for_month(month, budget)
    actual_total = df_expense['amount'].sum()
    actual_by_cat = df_expense.groupby('category')['amount'].sum().to_dict()
    
    over = {}
    # 检查总预算
    if '总预算' in month_budget and actual_total > month_budget['总预算']:
        over['总预算'] = actual_total - month_budget['总预算']
    
    # 检查各类别
    for cat, budget_val in month_budget.items():
        if cat == '总预算':
            continue
        actual = actual_by_cat.get(cat, 0)
        if actual > budget_val:
            over[cat] = actual - budget_val
    return over

def get_overbudget_details(month, df_month, budget):
    """
    返回超支详情字典列表，每个元素包含类别、预算、实际、超出
    """
    month_budget = get_budget_for_month(month, budget)
    if not month_budget:
        return []
    
    actual_by_cat = df_month.groupby('category')['amount'].sum().to_dict()
    details = []
    
    # 总预算超支检查
    if '总预算' in month_budget:
        total_actual = df_month['amount'].sum()
        total_budget = month_budget['总预算']
        if total_actual > total_budget:
            details.append({
                '月份': month,
                '类别': '总预算',
                '预算金额': total_budget,
                '实际支出': total_actual,
                '超出金额': total_actual - total_budget
            })
    
    # 各类别检查（排除总预算键）
    for cat, budget_val in month_budget.items():
        if cat == '总预算':
            continue
        actual = actual_by_cat.get(cat, 0)
        if actual > budget_val:
            details.append({
                '月份': month,
                '类别': cat,
                '预算金额': budget_val,
                '实际支出': actual,
                '超出金额': actual - budget_val
            })
    return details

def mark_excel_overbudget(excel_path, over_items, month):
    """
    在Excel报表的'月度汇总'表中标红超支的单元格
    over_items: {类别: 超出金额}
    """
    wb = load_workbook(excel_path)
    if '月度汇总' not in wb.sheetnames:
        return
    ws = wb['月度汇总']
    
    # 查找月份所在行，以及类别所在列
    # 假设第一列是月份，第一行是类别
    header = [cell.value for cell in ws[1]]
    month_col = 1  # 月份在第一列
    try:
        month_row = None
        for row in range(2, ws.max_row+1):
            if ws.cell(row, month_col).value == month:
                month_row = row
                break
        if month_row is None:
            return
        
        # 对每个超支类别，找到对应列标红
        for cat, _ in over_items.items():
            if cat == '总预算':
                # 总计列通常是最后一列
                total_col = len(header)
                ws.cell(month_row, total_col).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            else:
                if cat in header:
                    col = header.index(cat) + 1
                    ws.cell(month_row, col).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    except Exception as e:
        print(f"标记超支单元格时出错: {e}")
    finally:
        wb.save(excel_path)

def save_overbudget_csv(over_items, month, output_dir='output'):
    """将超支明细保存为CSV（追加模式）"""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    csv_path = output_path / 'overbudget_records.csv'
    
    rows = []
    for cat, exceed in over_items.items():
        rows.append({'月份': month, '类别': cat, '超出金额': exceed})
    
    df_new = pd.DataFrame(rows)
    if csv_path.exists():
        df_old = pd.read_csv(csv_path)
        df_all = pd.concat([df_old, df_new], ignore_index=True)
    else:
        df_all = df_new
    df_all.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"超支记录已保存到: {csv_path}")

def alert_all(report_generator, budget_path='config/budget.json'):
    """
    整合超支检查，传入report.py的生成器
    收集所有超支详情，保存到output/overbudget_details.csv，并在Excel中标记
    """
    budget = load_budget(budget_path)
    all_over_details = []  # 收集所有超支详情
    
    for excel_path, df_expense in report_generator:
        months = df_expense['month'].unique()
        for month in months:
            df_month = df_expense[df_expense['month'] == month]
            details = get_overbudget_details(month, df_month, budget)
            if details:
                print(f"⚠️ 超支提醒 ({month}):")
                for d in details:
                    print(f"   {d['类别']} 超出 {d['超出金额']:.2f} 元")
                all_over_details.extend(details)
                # 标记Excel单元格（转为{类别:超出金额}格式）
                over_dict = {d['类别']: d['超出金额'] for d in details}
                mark_excel_overbudget(excel_path, over_dict, month)
            else:
                print(f"✅ {month} 预算内")
    
    # 保存所有超支详情到CSV
    if all_over_details:
        df_all = pd.DataFrame(all_over_details)
        output_dir = Path('output')
        output_dir.mkdir(exist_ok=True)
        details_csv = output_dir / 'overbudget_details.csv'
        df_all.to_csv(details_csv, index=False, encoding='utf-8-sig')
        print(f"详细超支记录已保存到: {details_csv}")

if __name__ == '__main__':
    # 单独运行测试，需要先有report_generator
    from report import generate_monthly_report
    alert_all(generate_monthly_report())