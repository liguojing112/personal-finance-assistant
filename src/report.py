"""
生成月度汇总报表（Excel），包含明细表、月度汇总表、饼图、趋势图、柱状图、超支条形图
新增功能：
- 饼图图例显示类别+金额
- 月度总支出柱状图
- 各类别累计超支金额条形图（基于overbudget_details.csv）
"""
import pandas as pd
import numpy as np
from pathlib import Path
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import io
import base64

plt.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文
plt.rcParams['axes.unicode_minus'] = False

def generate_monthly_report(processed_dir='data/processed', output_dir='output'):
    """为每个清洗文件生成月度报表"""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    for parquet_file in Path(processed_dir).glob('*_cleaned.parquet'):
        print(f"生成报表: {parquet_file}")
        df = pd.read_parquet(parquet_file)
        
        # 确保只有支出用于汇总
        df_expense = df[df['income_expense'] == '支出'].copy()
        
        # 提取月份列
        df_expense['month'] = df_expense['tx_time'].dt.to_period('M').astype(str)
        
        # 明细表（添加类别列后）
        detail_sheet = df_expense[['tx_time', 'tx_type', 'counterparty', 'product', 'amount', 'category', 'remarks']]
        
        # 月度汇总表：按月份和类别统计支出总额
        monthly_summary = df_expense.groupby(['month', 'category'])['amount'].sum().unstack(fill_value=0)
        monthly_summary['总计'] = monthly_summary.sum(axis=1)
        
        # 总支出趋势（按月）
        trend = df_expense.groupby('month')['amount'].sum()
        
        # 创建Excel写入器
        base_name = parquet_file.stem.replace('_cleaned', '')
        excel_path = output_path / f"{base_name}_report.xlsx"
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # 写入明细表
            detail_sheet.to_excel(writer, sheet_name='明细', index=False)
            
            # 写入月度汇总表
            monthly_summary.to_excel(writer, sheet_name='月度汇总')
            
            # ---------- 饼图：各类别总支出占比（图例显示金额） ----------
            fig1, ax1 = plt.subplots(figsize=(6, 6))
            total_by_category = df_expense.groupby('category')['amount'].sum()
            # 过滤掉金额为0的类别，避免显示空类别
            total_by_category = total_by_category[total_by_category > 0]
            
            # 准备图例标签：类别 + 金额（元）
            legend_labels = [f"{cat}: {val:.0f}元" for cat, val in total_by_category.items()]
            
            # 画饼图，不自动显示标签（通过图例展示）
            wedges, texts, autotexts = ax1.pie(
                total_by_category.values,
                autopct='%1.1f%%',      # 仍显示百分比
                startangle=90,
                pctdistance=0.85        # 百分比文字距离圆心距离
            )
            # 设置图例，显示金额
            ax1.legend(wedges, legend_labels, title="支出类别", loc='center left', bbox_to_anchor=(1, 0, 0.5, 1))
            ax1.set_title('各类别支出占比（含金额）')
            
            # 保存饼图为字节流
            img_data_pie = io.BytesIO()
            plt.savefig(img_data_pie, format='png', dpi=100, bbox_inches='tight')
            plt.close(fig1)
            
            # ---------- 折线图：月度总支出趋势（原有） ----------
            fig2, ax2 = plt.subplots(figsize=(8, 4))
            trend.plot(kind='line', marker='o', ax=ax2)
            ax2.set_xlabel('月份')
            ax2.set_ylabel('支出金额(元)')
            ax2.set_title('月度总支出趋势')
            ax2.grid(True)
            img_data_trend = io.BytesIO()
            plt.savefig(img_data_trend, format='png', dpi=100, bbox_inches='tight')
            plt.close(fig2)
            
            # ---------- 新增柱状图：月度总支出柱状图 ----------
            fig3, ax3 = plt.subplots(figsize=(8, 4))
            # 提取月份列表和对应的金额
            months = trend.index.tolist()
            amounts = trend.values.tolist()
            ax3.bar(months, amounts, color='skyblue')
            ax3.set_xlabel('月份')
            ax3.set_ylabel('支出金额(元)')
            ax3.set_title('月度总支出柱状图')
            # 旋转x轴标签避免重叠
            plt.xticks(rotation=45)
            # 在柱子上方标注数值
            for i, v in enumerate(amounts):
                ax3.text(i, v + (max(amounts)*0.02), f'{v:.0f}', ha='center', va='bottom')
            ax3.grid(axis='y', linestyle='--', alpha=0.7)
            plt.tight_layout()
            img_data_bar = io.BytesIO()
            plt.savefig(img_data_bar, format='png', dpi=100, bbox_inches='tight')
            plt.close(fig3)
            
            # ---------- 读取超支明细并绘制超支条形图 ----------
            over_csv = output_path / 'overbudget_details.csv'
            img_data_over = None
            if over_csv.exists():
                df_over = pd.read_csv(over_csv)
                if not df_over.empty:
                    # 计算各类别累计超支金额（排除总预算类别？可以保留）
                    cat_over_sum = df_over.groupby('类别')['超出金额'].sum().sort_values(ascending=False)
                    if not cat_over_sum.empty:
                        fig4, ax4 = plt.subplots(figsize=(8, 4))
                        cat_over_sum.plot(kind='bar', ax=ax4, color='salmon')
                        ax4.set_xlabel('支出类别')
                        ax4.set_ylabel('超支金额(元)')
                        ax4.set_title('各类别累计超支金额')
                        ax4.grid(axis='y', linestyle='--', alpha=0.7)
                        # 在柱顶标注数值
                        max_val = max(cat_over_sum.values) if len(cat_over_sum) > 0 else 1
                        for i, v in enumerate(cat_over_sum.values):
                            ax4.text(i, v + max_val*0.02, f'{v:.0f}', ha='center', va='bottom')
                        plt.tight_layout()
                        img_data_over = io.BytesIO()
                        plt.savefig(img_data_over, format='png', dpi=100, bbox_inches='tight')
                        plt.close(fig4)
            
            # ---------- 将所有图片插入到“图表”工作表 ----------
            workbook = writer.book
            if '图表' in workbook.sheetnames:
                std = workbook['图表']
                workbook.remove(std)
            chart_sheet = workbook.create_sheet('图表')
            
            # 插入饼图（位置A1）
            img_pie = Image(img_data_pie)
            img_pie.anchor = 'A1'
            chart_sheet.add_image(img_pie)
            
            # 插入折线图（位置A20）
            img_trend = Image(img_data_trend)
            img_trend.anchor = 'A20'
            chart_sheet.add_image(img_trend)
            
            # 插入柱状图（位置A40）
            img_bar = Image(img_data_bar)
            img_bar.anchor = 'A40'
            chart_sheet.add_image(img_bar)
            
            # 插入超支条形图（位置A60，如果存在）
            if img_data_over:
                img_over = Image(img_data_over)
                img_over.anchor = 'A60'
                chart_sheet.add_image(img_over)
        
        print(f"报表已生成: {excel_path}")
        
        # 为后续超支提醒模块返回报表路径和df_expense
        yield excel_path, df_expense

if __name__ == '__main__':
    for _ in generate_monthly_report():
        pass

def generate_charts_base64(df_expense):
    """从支出DataFrame生成饼图和趋势图的Base64编码字符串"""
    import io
    import base64
    import matplotlib.pyplot as plt

    # 确保只有支出
    df_expense = df_expense[df_expense['income_expense'] == '支出'].copy()
    if df_expense.empty:
        return None, None

    # 提取月份列
    df_expense['month'] = df_expense['tx_time'].dt.to_period('M').astype(str)
    
    # 1. 饼图
    fig1, ax1 = plt.subplots(figsize=(6, 6))
    total_by_category = df_expense.groupby('category')['amount'].sum()
    total_by_category = total_by_category[total_by_category > 0]
    legend_labels = [f"{cat}: {val:.0f}元" for cat, val in total_by_category.items()]
    wedges, texts, autotexts = ax1.pie(
        total_by_category.values,
        autopct='%1.1f%%',
        startangle=90,
        pctdistance=0.85
    )
    ax1.legend(wedges, legend_labels, title="支出类别", loc='center left', bbox_to_anchor=(1, 0, 0.5, 1))
    ax1.set_title('各类别支出占比（含金额）')
    plt.tight_layout()
    
    img_pie = io.BytesIO()
    plt.savefig(img_pie, format='png', dpi=100, bbox_inches='tight')
    plt.close(fig1)
    img_pie.seek(0)
    pie_base64 = base64.b64encode(img_pie.getvalue()).decode('utf-8')

    # 2. 趋势图（折线图）
    trend = df_expense.groupby('month')['amount'].sum()
    fig2, ax2 = plt.subplots(figsize=(8, 4))
    trend.plot(kind='line', marker='o', ax=ax2)
    ax2.set_xlabel('月份')
    ax2.set_ylabel('支出金额(元)')
    ax2.set_title('月度总支出趋势')
    ax2.grid(True)
    plt.tight_layout()
    
    img_trend = io.BytesIO()
    plt.savefig(img_trend, format='png', dpi=100, bbox_inches='tight')
    plt.close(fig2)
    img_trend.seek(0)
    trend_base64 = base64.b64encode(img_trend.getvalue()).decode('utf-8')

    return pie_base64, trend_base64


def plot_pie_chart(df_expense):
    """生成饼图，返回base64编码的图片字符串"""
    total_by_category = df_expense.groupby('category')['amount'].sum()
    total_by_category = total_by_category[total_by_category > 0]  # 过滤掉金额为0的类别
    
    fig, ax = plt.subplots(figsize=(6, 6))
    wedges, texts, autotexts = ax.pie(
        total_by_category.values,
        autopct='%1.1f%%',
        startangle=90,
        pctdistance=0.85
    )
    legend_labels = [f"{cat}: {val:.0f}元" for cat, val in total_by_category.items()]
    ax.legend(wedges, legend_labels, title="支出类别", loc='center left', bbox_to_anchor=(1, 0, 0.5, 1))
    ax.set_title('各类别支出占比（含金额）')
    
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    img_base64 = base64.b64encode(buf.getvalue()).decode('utf-8')
    return img_base64

def plot_trend_chart(df_expense):
    """生成月度趋势折线图，返回base64编码的图片字符串"""
    # 提取月份和总支出
    trend = df_expense.groupby(df_expense['tx_time'].dt.to_period('M'))['amount'].sum()
    months = trend.index.astype(str).tolist()
    amounts = trend.values.tolist()
    
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.plot(months, amounts, marker='o', linewidth=2, markersize=8)
    ax.set_xlabel('月份')
    ax.set_ylabel('支出金额(元)')
    ax.set_title('月度总支出趋势')
    ax.grid(True, linestyle='--', alpha=0.7)
    
    # 在点上标注数值
    for i, v in enumerate(amounts):
        ax.text(i, v + max(amounts)*0.02, f'{v:.0f}', ha='center', va='bottom')
    
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    img_base64 = base64.b64encode(buf.getvalue()).decode('utf-8')
    return img_base64